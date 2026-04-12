from __future__ import annotations

import hashlib
import re
from pathlib import Path
from threading import Lock

from openpyxl import load_workbook

from .....config import AppSettings
from .db_repository import ensure_eco_fee_tables, read_eco_fee_catalog, replace_eco_fee_catalog
from .models import (
    EcoFeeCatalog,
    EcoFeeCatalogDbLoadResult,
    EcoFeeCatalogDbSyncResult,
    EcoFeeWorkbookParseResult,
    EcoGroupYearValue,
    EcoMapEntry,
)


GROUP_CODE_RE = re.compile(r"Группа\s*N\s*(\d+)", re.IGNORECASE)
YEAR_RE = re.compile(r"20\d{2}")
FOOTNOTE_RE = re.compile(r"<\s*(\d+)\s*>")
EXPLICIT_CODE_RE = re.compile(r"\d{4,10}")
PACKAGE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK_PATH = PACKAGE_DIR / "Расчет экосбора.xlsx"
SCIENTIFIC_CODE_OVERRIDES: dict[str, tuple[str, ...]] = {
    "4.82361000048236e+19": ("4823610000", "4823699000"),
    "8.51631000851632e+18": ("8516310001", "8516310009", "8516320000"),
    "9.10121000091012e+29": ("9101210000", "9101290000"),
    "9.10111000091011e+29": ("9101110000", "9101190000"),
    "9.40531000094053e+19": ("9405310000", "9405390000"),
    "3.92330101039233e+19": ("3923301010", "3923309010"),
}
ALLOWED_TNVED_LENGTHS = {4, 6, 7, 8, 9, 10}


class EcoFeeKnowledgeCatalogService:
    def __init__(
        self,
        *,
        settings: AppSettings | None = None,
        workbook_path: str | Path | None = None,
    ) -> None:
        self._settings = settings
        self._workbook_path = Path(workbook_path).expanduser().resolve() if workbook_path else DEFAULT_WORKBOOK_PATH
        self._lock = Lock()
        self._cache: EcoFeeCatalog | None = None
        self._cache_source: str = ""

    @staticmethod
    def _cell_text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            text = f"{value:.8f}".rstrip("0").rstrip(".")
            return text or "0"
        return str(value).strip()

    @staticmethod
    def _to_float(value: object) -> float | None:
        text = EcoFeeKnowledgeCatalogService._cell_text(value).replace(" ", "").replace(",", ".")
        if not text or text in {"-", "—"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _normalize_code(value: object) -> str:
        text = EcoFeeKnowledgeCatalogService._cell_text(value)
        # In the workbook TN VED codes often carry inline footnotes like `6301<3>`.
        # For matching we need the code itself, not the footnote marker.
        text = FOOTNOTE_RE.sub("", text)
        return re.sub(r"\D+", "", text)

    @staticmethod
    def _normalize_codes(value: object) -> tuple[str, ...]:
        if isinstance(value, float):
            scientific_text = str(value).strip()
            if scientific_text in SCIENTIFIC_CODE_OVERRIDES:
                return SCIENTIFIC_CODE_OVERRIDES[scientific_text]

        text = EcoFeeKnowledgeCatalogService._cell_text(value)
        if not text:
            return ()

        if text in SCIENTIFIC_CODE_OVERRIDES:
            return SCIENTIFIC_CODE_OVERRIDES[text]

        explicit_codes = [item for item in EXPLICIT_CODE_RE.findall(text) if 4 <= len(item) <= 10]
        if explicit_codes:
            unique_codes: list[str] = []
            for code in explicit_codes:
                if code not in unique_codes:
                    unique_codes.append(code)
            return tuple(unique_codes)

        text_without_footnotes = FOOTNOTE_RE.sub("", text)
        digits = re.sub(r"\D+", "", text_without_footnotes)
        if 4 <= len(digits) <= 10:
            return (digits,)
        return ()

    @staticmethod
    def _extract_footnote_refs(*values: object) -> tuple[str, ...]:
        refs: list[str] = []
        for value in values:
            text = EcoFeeKnowledgeCatalogService._cell_text(value)
            for ref in FOOTNOTE_RE.findall(text):
                if ref not in refs:
                    refs.append(ref)
        return tuple(refs)

    @staticmethod
    def _parse_group_code(group_name: str) -> str | None:
        match = GROUP_CODE_RE.search(group_name)
        return match.group(1) if match else None

    @staticmethod
    def _sheet_by_name_part(workbook, marker: str) -> object:
        for sheet_name in workbook.sheetnames:
            if marker in sheet_name:
                return workbook[sheet_name]
        raise KeyError(f"Worksheet with marker '{marker}' not found.")

    @staticmethod
    def _sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def _parse_rates(self, workbook) -> dict[int, dict[str, EcoGroupYearValue]]:
        sheet = self._sheet_by_name_part(workbook, "2025-2027")
        header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        year_columns: dict[int, int] = {}
        for index, value in enumerate(header_row, start=1):
            text = self._cell_text(value)
            match = YEAR_RE.search(text)
            if match:
                year_columns[int(match.group(0))] = index

        groups_by_year: dict[int, dict[str, EcoGroupYearValue]] = {year: {} for year in year_columns}
        for row in sheet.iter_rows(min_row=3, values_only=True):
            group_name = self._cell_text(row[0] if len(row) >= 1 else "")
            group_code = self._parse_group_code(group_name)
            if not group_code:
                continue
            for year, column_index in year_columns.items():
                rate_ton = self._to_float(row[column_index - 1] if len(row) >= column_index else None)
                groups_by_year[year][group_code] = EcoGroupYearValue(
                    eco_group_code=group_code,
                    eco_group_name=group_name,
                    year=year,
                    rate_rub_per_ton=rate_ton,
                    rate_rub_per_kg=(rate_ton / 1000.0) if rate_ton is not None else None,
                    complexity_coeff=None,
                    utilization_norm=None,
                )
        return groups_by_year

    def _merge_coefficients(self, workbook, groups_by_year: dict[int, dict[str, EcoGroupYearValue]]) -> None:
        for year in sorted(groups_by_year):
            try:
                sheet = self._sheet_by_name_part(workbook, f"извлечения {year}")
            except KeyError:
                continue
            for row in sheet.iter_rows(min_row=3, values_only=True):
                group_name = self._cell_text(row[0] if len(row) >= 1 else "")
                group_code = self._parse_group_code(group_name)
                if not group_code or year not in groups_by_year or group_code not in groups_by_year[year]:
                    continue
                current = groups_by_year[year][group_code]
                groups_by_year[year][group_code] = EcoGroupYearValue(
                    eco_group_code=current.eco_group_code,
                    eco_group_name=current.eco_group_name,
                    year=current.year,
                    rate_rub_per_ton=current.rate_rub_per_ton,
                    rate_rub_per_kg=current.rate_rub_per_kg,
                    complexity_coeff=self._to_float(row[1] if len(row) >= 2 else None),
                    utilization_norm=current.utilization_norm,
                )

    def _merge_goods_norms(self, workbook, groups_by_year: dict[int, dict[str, EcoGroupYearValue]]) -> None:
        sheet = self._sheet_by_name_part(workbook, "товаров")
        header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        year_columns: dict[int, int] = {}
        for index, value in enumerate(header_row, start=1):
            text = self._cell_text(value)
            match = YEAR_RE.search(text)
            if match:
                year_columns[int(match.group(0))] = index

        for row in sheet.iter_rows(min_row=3, values_only=True):
            group_name = self._cell_text(row[0] if len(row) >= 1 else "")
            group_code = self._parse_group_code(group_name)
            if not group_code:
                continue
            for year, column_index in year_columns.items():
                if year not in groups_by_year or group_code not in groups_by_year[year]:
                    continue
                norm_percent = self._to_float(row[column_index - 1] if len(row) >= column_index else None)
                current = groups_by_year[year][group_code]
                groups_by_year[year][group_code] = EcoGroupYearValue(
                    eco_group_code=current.eco_group_code,
                    eco_group_name=current.eco_group_name,
                    year=current.year,
                    rate_rub_per_ton=current.rate_rub_per_ton,
                    rate_rub_per_kg=current.rate_rub_per_kg,
                    complexity_coeff=current.complexity_coeff,
                    utilization_norm=(norm_percent / 100.0) if norm_percent is not None else None,
                )

    def _parse_packaging_norms(self, workbook) -> dict[int, float]:
        sheet = self._sheet_by_name_part(workbook, "упаковки")
        notes = " ".join(
            self._cell_text(row[0])
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True)
            if row and row[0]
        )
        norms: dict[int, float] = {}
        for raw_year in sorted({int(match.group(0)) for match in YEAR_RE.finditer(notes)}):
            year = int(raw_year)
            decimal_match = re.search(rf"{year}[^\d]{{0,40}}(\d+[.,]\d+)", notes)
            if decimal_match:
                norms[year] = float(decimal_match.group(1).replace(",", "."))
                continue
            percent_match = re.search(rf"{year}[^\d]{{0,40}}(\d+(?:[.,]\d+)?)\s*%", notes)
            if percent_match:
                norms[year] = float(percent_match.group(1).replace(",", ".")) / 100.0
        return norms

    def _parse_usd_rate(self, workbook) -> float | None:
        sheet = self._sheet_by_name_part(workbook, "Курсы валют")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code = self._cell_text(row[0] if len(row) >= 1 else "")
            if code.upper() == "USD":
                return self._to_float(row[1] if len(row) >= 2 else None)
        return None

    def _parse_map_entries(self, workbook) -> tuple[EcoMapEntry, ...]:
        sheet = self._sheet_by_name_part(workbook, "2414")
        collected: list[EcoMapEntry] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            row_name = self._cell_text(row[0] if len(row) >= 1 else "")
            okpd2 = self._cell_text(row[1] if len(row) >= 2 else "")
            tnved_raw = self._cell_text(row[2] if len(row) >= 3 else "")
            tnved_codes = self._normalize_codes(row[2] if len(row) >= 3 else None)
            tnved_name = self._cell_text(row[3] if len(row) >= 4 else "")
            eco_group_name = self._cell_text(row[4] if len(row) >= 5 else "")
            eco_group_code = self._parse_group_code(eco_group_name)
            display_name = row_name or tnved_name
            if not display_name or not tnved_codes or not eco_group_code:
                continue
            footnote_refs = self._extract_footnote_refs(row_name, tnved_raw, tnved_name)
            for tnved_digits in tnved_codes:
                if len(tnved_digits) not in ALLOWED_TNVED_LENGTHS:
                    continue
                collected.append(
                    EcoMapEntry(
                        source_row=row_index,
                        row_name=display_name,
                        okpd2=okpd2,
                        tnved_raw=tnved_raw,
                        tnved_digits=tnved_digits,
                        tnved_name=tnved_name,
                        eco_group_code=eco_group_code,
                        eco_group_name=eco_group_name,
                        footnote_refs=footnote_refs,
                    )
                )
        return tuple(collected)

    def _parse_footnotes(self, workbook) -> dict[str, str]:
        sheet = self._sheet_by_name_part(workbook, "2414")
        footnotes: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=1, values_only=True):
            row_name = self._cell_text(row[0] if len(row) >= 1 else "")
            match = FOOTNOTE_RE.match(row_name)
            if not match:
                continue
            ref = match.group(1)
            text = FOOTNOTE_RE.sub("", row_name, count=1).strip(" .;-")
            if text and ref not in footnotes:
                footnotes[ref] = text
        return footnotes

    def parse_workbook(self, workbook_path: str | Path | None = None) -> tuple[EcoFeeCatalog, EcoFeeWorkbookParseResult]:
        resolved_path = Path(workbook_path).expanduser().resolve() if workbook_path else self._workbook_path
        if not resolved_path.exists():
            raise FileNotFoundError(f"Eco fee workbook not found: {resolved_path}")

        workbook = load_workbook(resolved_path, data_only=True, read_only=True)
        try:
            groups_by_year = self._parse_rates(workbook)
            self._merge_coefficients(workbook, groups_by_year)
            self._merge_goods_norms(workbook, groups_by_year)
            supported_years = tuple(sorted(groups_by_year))
            default_year = 2026 if 2026 in supported_years else (supported_years[0] if supported_years else 2026)
            catalog = EcoFeeCatalog(
                supported_years=supported_years,
                default_year=default_year,
                usd_rate=self._parse_usd_rate(workbook),
                packaging_norms=self._parse_packaging_norms(workbook),
                groups_by_year=groups_by_year,
                map_entries=self._parse_map_entries(workbook),
                footnotes=self._parse_footnotes(workbook),
            )
            result = EcoFeeWorkbookParseResult(
                workbook_path=str(resolved_path),
                source_hash=self._sha256(resolved_path),
                sheet_names=tuple(workbook.sheetnames),
                supported_years=supported_years,
                groups_count=sum(len(groups) for groups in groups_by_year.values()),
                map_entries_count=len(catalog.map_entries),
            )
            return catalog, result
        finally:
            workbook.close()

    def sync_to_postgres(
        self,
        *,
        workbook_path: str | Path | None = None,
        database_url: str | None = None,
    ) -> EcoFeeCatalogDbSyncResult:
        catalog, parse_result = self.parse_workbook(workbook_path=workbook_path)
        ensure_eco_fee_tables(database_url=database_url)
        return replace_eco_fee_catalog(
            catalog=catalog,
            source_path=parse_result.workbook_path,
            source_hash=parse_result.source_hash,
            sheet_names=parse_result.sheet_names,
            database_url=database_url,
        )

    def build_runtime_catalog(
        self,
        *,
        prefer_database: bool = True,
        database_url: str | None = None,
    ) -> tuple[EcoFeeCatalog, dict[str, object]]:
        if prefer_database:
            try:
                ensure_eco_fee_tables(database_url=database_url)
                catalog, load_result = read_eco_fee_catalog(database_url=database_url)
                if (catalog.map_entries or catalog.groups_by_year) and catalog.footnotes:
                    return catalog, {"source": "postgres", "load": load_result.to_payload()}
                if catalog.map_entries or catalog.groups_by_year:
                    fallback_error = "postgres catalog is missing eco footnotes; fallback to workbook"
                else:
                    fallback_error = ""
            except Exception as exc:
                fallback_error = str(exc)
        else:
            fallback_error = ""

        catalog, parse_result = self.parse_workbook()
        return catalog, {
            "source": "workbook",
            "load": parse_result.to_payload(),
            "fallback_error": fallback_error,
        }

    def get_catalog(
        self,
        *,
        prefer_database: bool = True,
        database_url: str | None = None,
    ) -> EcoFeeCatalog:
        cache_key = f"{int(prefer_database)}:{database_url or ''}:{self._workbook_path}"
        with self._lock:
            if self._cache is None or self._cache_source != cache_key:
                self._cache, _ = self.build_runtime_catalog(
                    prefer_database=prefer_database,
                    database_url=database_url,
                )
                self._cache_source = cache_key
            return self._cache


__all__ = [
    "DEFAULT_WORKBOOK_PATH",
    "EcoFeeKnowledgeCatalogService",
]
