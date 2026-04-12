from __future__ import annotations

import io
import json
import re
import shutil
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..repository import TnvedCatalogSnapshot, build_tnved_catalog_snapshot, normalize_code_10
from .db_repository import (
    TnvedCatalogDbLoadResult,
    TnvedCatalogDbSyncResult,
    ensure_tnved_catalog_tables,
    read_tnved_catalog_snapshot,
    replace_tnved_catalog,
)
from .models import (
    TnvedCatalogRecord,
    TnvedCatalogState,
    TnvedCatalogUpdateResult,
    TnvedDownloadResult,
    TnvedWorkbookParseResult,
)


DEFAULT_TNVED_DOWNLOAD_URL = "https://www.tws.by/tws/tnved/download/excel"
DEFAULT_TNVED_SHEET_NAME = "ТНВЭД"
PACKAGE_DIR = Path(__file__).resolve().parent
STATE_JSON_PATH = PACKAGE_DIR / "state.json"
CATALOG_JSON_PATH = PACKAGE_DIR / "catalog.json"
DOWNLOADS_DIR = PACKAGE_DIR / "downloads"
BACKUPS_DIR = PACKAGE_DIR / "backups"

_CODE_HEADER_ALIASES = frozenset({"код", "код тнвэд", "код тн вэд", "тнвэд", "тн вэд", "tnved", "code"})
_NAME_HEADER_ALIASES = frozenset({"наименование", "описание", "товар", "description", "name"})
_DUTY_HEADER_ALIASES = frozenset(
    {"тариф", "пошлина", "ставка пошлины", "ставка ввозной пошлины", "duty", "duty rate", "tariff"}
)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _timestamp_label() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("\xa0", " ")
    return re.sub(r"\s+", " ", text)


def _resolve_columns(header_row: tuple[object, ...]) -> tuple[int, int, int | None]:
    headers = {_normalize_header(value): idx for idx, value in enumerate(header_row)}
    code_idx = next((idx for name, idx in headers.items() if name in _CODE_HEADER_ALIASES), None)
    name_idx = next((idx for name, idx in headers.items() if name in _NAME_HEADER_ALIASES), None)
    duty_idx = next((idx for name, idx in headers.items() if name in _DUTY_HEADER_ALIASES), None)
    if code_idx is None or name_idx is None:
        raise RuntimeError("В листе должны быть колонки минимум: 'Код' и 'Наименование'")
    return code_idx, name_idx, duty_idx


def _worksheet_from_workbook(workbook: Workbook, preferred_sheet_name: str | None) -> Worksheet:
    if preferred_sheet_name and preferred_sheet_name in workbook.sheetnames:
        return workbook[preferred_sheet_name]
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        try:
            _resolve_columns(header_row)
            return worksheet
        except RuntimeError:
            continue
    raise RuntimeError("Не найден лист с колонками 'Код' и 'Наименование'")


def parse_tnved_workbook(
    xlsx_path: str | Path,
    *,
    preferred_sheet_name: str | None = DEFAULT_TNVED_SHEET_NAME,
) -> TnvedWorkbookParseResult:
    resolved_path = Path(xlsx_path).expanduser().resolve()
    if not resolved_path.exists():
        raise FileNotFoundError(f"Excel file not found: {resolved_path}")

    workbook = load_workbook(io.BytesIO(resolved_path.read_bytes()), read_only=True, data_only=True)
    try:
        worksheet = _worksheet_from_workbook(workbook, preferred_sheet_name)
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise RuntimeError("Лист пустой, заголовки не найдены")

        code_idx, name_idx, duty_idx = _resolve_columns(header_row)
        scanned_rows = 0
        records_map: dict[str, TnvedCatalogRecord] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            scanned_rows += 1
            code = normalize_code_10(row[code_idx] if code_idx < len(row) else None)
            if not code:
                continue
            description = str(row[name_idx] if name_idx < len(row) else "" or "").strip()
            if not description:
                continue
            duty_rate_raw = str(row[duty_idx] if duty_idx is not None and duty_idx < len(row) else "" or "").strip()
            records_map[code] = TnvedCatalogRecord(code=code, description=description, duty_rate=duty_rate_raw or None)
        records = tuple(records_map.values())
        if not records:
            raise RuntimeError("Нет валидных строк для импорта")
        return TnvedWorkbookParseResult(
            workbook_path=str(resolved_path),
            sheet_name=worksheet.title,
            scanned_rows=scanned_rows,
            valid_unique_records=len(records),
            records=records,
        )
    finally:
        workbook.close()


def download_tnved_workbook(
    *,
    source_url: str = DEFAULT_TNVED_DOWNLOAD_URL,
    destination_path: str | Path,
    timeout_sec: int = 60,
) -> TnvedDownloadResult:
    resolved_destination = Path(destination_path).expanduser().resolve()
    resolved_destination.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(
        source_url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) agent_ui tnved updater",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout_sec) as response:
        payload = response.read()
        content_type = response.headers.get("Content-Type")
    if not payload.startswith(b"PK\x03\x04"):
        raise RuntimeError("Downloaded payload is not a valid .xlsx archive")
    resolved_destination.write_bytes(payload)
    return TnvedDownloadResult(
        source_url=source_url,
        destination_path=str(resolved_destination),
        size_bytes=len(payload),
        content_type=content_type,
    )


class TnvedCatalogService:
    def __init__(
        self,
        *,
        state_json_path: str | Path | None = STATE_JSON_PATH,
        catalog_json_path: str | Path | None = CATALOG_JSON_PATH,
        downloads_dir: str | Path | None = DOWNLOADS_DIR,
        backups_dir: str | Path | None = BACKUPS_DIR,
    ) -> None:
        self._state_json_path = Path(state_json_path or STATE_JSON_PATH).expanduser().resolve()
        self._catalog_json_path = Path(catalog_json_path or CATALOG_JSON_PATH).expanduser().resolve()
        self._downloads_dir = Path(downloads_dir or DOWNLOADS_DIR).expanduser().resolve()
        self._backups_dir = Path(backups_dir or BACKUPS_DIR).expanduser().resolve()
        self._ensure_layout()

    def _ensure_layout(self) -> None:
        self._state_json_path.parent.mkdir(parents=True, exist_ok=True)
        self._catalog_json_path.parent.mkdir(parents=True, exist_ok=True)
        self._downloads_dir.mkdir(parents=True, exist_ok=True)
        self._backups_dir.mkdir(parents=True, exist_ok=True)
        if not self._state_json_path.exists():
            self.write_state(TnvedCatalogState())
        if not self._catalog_json_path.exists():
            self._catalog_json_path.write_text(
                json.dumps(
                    {
                        "status": "empty",
                        "updated_at": "",
                        "source_url": "",
                        "workbook_path": "",
                        "sheet_name": "",
                        "records": [],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

    def read_state(self) -> TnvedCatalogState:
        self._ensure_layout()
        payload = json.loads(self._state_json_path.read_text(encoding="utf-8"))
        return TnvedCatalogState(
            status=str(payload.get("status") or "empty"),
            active_catalog_path=str(payload.get("active_catalog_path") or ""),
            backup_catalog_path=str(payload.get("backup_catalog_path") or ""),
            source_url=str(payload.get("source_url") or ""),
            workbook_path=str(payload.get("workbook_path") or ""),
            sheet_name=str(payload.get("sheet_name") or ""),
            active_rows=int(payload.get("active_rows") or 0),
            scanned_rows=int(payload.get("scanned_rows") or 0),
            imported_rows=int(payload.get("imported_rows") or 0),
            updated_at=str(payload.get("updated_at") or ""),
            notes=tuple(str(item) for item in (payload.get("notes") or []) if str(item).strip()),
        )

    def write_state(self, state: TnvedCatalogState) -> None:
        self._state_json_path.write_text(json.dumps(state.to_payload(), ensure_ascii=False, indent=2), encoding="utf-8")

    def read_catalog_payload(self) -> dict[str, Any]:
        self._ensure_layout()
        return json.loads(self._catalog_json_path.read_text(encoding="utf-8"))

    def build_snapshot(self) -> TnvedCatalogSnapshot:
        payload = self.read_catalog_payload()
        records = [
            (
                str(item.get("code") or ""),
                str(item.get("description") or ""),
                str(item.get("duty_rate") or "") or None,
            )
            for item in (payload.get("records") or [])
            if isinstance(item, dict)
        ]
        return build_tnved_catalog_snapshot(records)

    def build_runtime_snapshot(
        self,
        *,
        prefer_database: bool = True,
        database_url: str | None = None,
    ) -> tuple[TnvedCatalogSnapshot | None, dict[str, Any]]:
        if prefer_database:
            try:
                snapshot, load_result = read_tnved_catalog_snapshot(database_url=database_url)
                if snapshot.codes_set:
                    return snapshot, {
                        "source": "postgres",
                        "load": load_result.to_payload(),
                    }
            except Exception as exc:
                fallback_error = str(exc)
            else:
                fallback_error = ""
        else:
            fallback_error = ""

        snapshot = self.build_snapshot()
        return snapshot, {
            "source": "json",
            "load": {
                "loaded_rows": len(snapshot.codes_set),
            },
            "fallback_error": fallback_error,
        }

    def count_rows(self) -> int:
        payload = self.read_catalog_payload()
        return len(payload.get("records") or [])

    def _backup_catalog_json(self) -> str:
        if not self._catalog_json_path.exists():
            return ""
        stamp = _timestamp_label()
        backup_path = self._backups_dir / f"catalog_{stamp}.json"
        shutil.copy2(self._catalog_json_path, backup_path)
        return str(backup_path)

    def update_from_workbook(
        self,
        *,
        workbook_path: str | Path,
        source_url: str = "",
        preferred_sheet_name: str | None = DEFAULT_TNVED_SHEET_NAME,
    ) -> TnvedCatalogUpdateResult:
        previous_rows = self.count_rows()
        parse_result = parse_tnved_workbook(workbook_path, preferred_sheet_name=preferred_sheet_name)
        backup_path = self._backup_catalog_json()
        catalog_payload = {
            "status": "ready",
            "updated_at": _utc_now_iso(),
            "source_url": source_url,
            "workbook_path": parse_result.workbook_path,
            "sheet_name": parse_result.sheet_name,
            "records": [record.to_payload() for record in parse_result.records],
        }
        self._catalog_json_path.write_text(json.dumps(catalog_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        notes = ("catalog_replaced_from_workbook",)
        state = TnvedCatalogState(
            status="ready",
            active_catalog_path=str(self._catalog_json_path),
            backup_catalog_path=backup_path,
            source_url=source_url,
            workbook_path=parse_result.workbook_path,
            sheet_name=parse_result.sheet_name,
            active_rows=parse_result.valid_unique_records,
            scanned_rows=parse_result.scanned_rows,
            imported_rows=parse_result.valid_unique_records,
            updated_at=catalog_payload["updated_at"],
            notes=notes,
        )
        self.write_state(state)
        return TnvedCatalogUpdateResult(
            status="ready",
            workbook_path=parse_result.workbook_path,
            catalog_json_path=str(self._catalog_json_path),
            state_json_path=str(self._state_json_path),
            backup_json_path=backup_path,
            sheet_name=parse_result.sheet_name,
            previous_rows=previous_rows,
            scanned_rows=parse_result.scanned_rows,
            imported_rows=parse_result.valid_unique_records,
            active_rows=parse_result.valid_unique_records,
            source_url=source_url,
            notes=notes,
        )

    def download_and_update(
        self,
        *,
        source_url: str = DEFAULT_TNVED_DOWNLOAD_URL,
        preferred_sheet_name: str | None = DEFAULT_TNVED_SHEET_NAME,
        timeout_sec: int = 60,
    ) -> TnvedCatalogUpdateResult:
        download_path = self._downloads_dir / f"tws_tnved_{_timestamp_label()}.xlsx"
        download_tnved_workbook(source_url=source_url, destination_path=download_path, timeout_sec=timeout_sec)
        return self.update_from_workbook(
            workbook_path=download_path,
            source_url=source_url,
            preferred_sheet_name=preferred_sheet_name,
        )

    def sync_to_postgres(self, *, database_url: str | None = None) -> TnvedCatalogDbSyncResult:
        payload = self.read_catalog_payload()
        records = tuple(
            TnvedCatalogRecord(
                code=str(item.get("code") or ""),
                description=str(item.get("description") or ""),
                duty_rate=str(item.get("duty_rate") or "") or None,
            )
            for item in (payload.get("records") or [])
            if isinstance(item, dict)
        )
        state = self.read_state()
        ensure_tnved_catalog_tables(database_url)
        return replace_tnved_catalog(records=records, state=state, database_url=database_url)

    def download_update_and_sync(
        self,
        *,
        source_url: str = DEFAULT_TNVED_DOWNLOAD_URL,
        preferred_sheet_name: str | None = DEFAULT_TNVED_SHEET_NAME,
        timeout_sec: int = 60,
        database_url: str | None = None,
    ) -> tuple[TnvedCatalogUpdateResult, TnvedCatalogDbSyncResult]:
        update_result = self.download_and_update(
            source_url=source_url,
            preferred_sheet_name=preferred_sheet_name,
            timeout_sec=timeout_sec,
        )
        db_result = self.sync_to_postgres(database_url=database_url)
        return update_result, db_result


__all__ = [
    "CATALOG_JSON_PATH",
    "DEFAULT_TNVED_DOWNLOAD_URL",
    "DEFAULT_TNVED_SHEET_NAME",
    "DOWNLOADS_DIR",
    "STATE_JSON_PATH",
    "TnvedCatalogDbLoadResult",
    "TnvedCatalogService",
    "download_tnved_workbook",
    "parse_tnved_workbook",
]
