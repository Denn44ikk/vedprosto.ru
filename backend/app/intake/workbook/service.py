from __future__ import annotations

import re
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook

from ...config import AppSettings
from ...orchestrator.job_store import JobStore
from ...storage.runtime_state.service import RuntimeStateService
from .exporter import collect_xlsx_images_by_row, ensure_openpyxl_workbook, export_workbook_cases
from .source_workbook_service import SourceWorkbookService


WORKBOOK_REQUIRED_HEADERS = ["Наименование"]
WORKBOOK_OPTIONAL_HEADERS = ["Доп информация", "Фото"]
WORKBOOK_PREVIEW_LIMIT = 250
DISPLAY_NAME_HEADERS = ("Наименование", "product_name", "Название", "Name")
SECONDARY_TEXT_HEADERS = ("Доп информация", "Доп инфа", "article", "manufacturer", "applicant", "row_text")
HEADER_NAME_ALIASES = {
    "Наименование": {
        "наименование",
        "product_name",
        "название",
        "name",
    },
    "Доп информация": {
        "доп информация",
        "доп. информация",
        "доп инфа",
        "доп. инфа",
        "additional info",
        "extra info",
        "description",
        "row_text",
    },
}


class WorkbookIntakeService:
    def __init__(
        self,
        *,
        settings: AppSettings,
        job_store: JobStore,
        runtime_state_service: RuntimeStateService,
        source_workbook_service: SourceWorkbookService,
    ) -> None:
        self.settings = settings
        self.job_store = job_store
        self.runtime_state_service = runtime_state_service
        self.source_workbook_service = source_workbook_service
        self.uploads_dir = settings.uploads_dir
        self.uploads_dir.mkdir(parents=True, exist_ok=True)

    async def save_upload(self, workbook_file: UploadFile) -> Path:
        target = self.uploads_dir / workbook_file.filename
        suffix = target.suffix
        stem = target.stem
        counter = 1
        while target.exists():
            target = self.uploads_dir / f"{stem}_{counter}{suffix}"
            counter += 1

        with target.open("wb") as stream:
            while True:
                chunk = await workbook_file.read(1024 * 1024)
                if not chunk:
                    break
                stream.write(chunk)
        await workbook_file.close()
        return target

    @staticmethod
    def _normalize_cell(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_header(value: object) -> str:
        text = WorkbookIntakeService._normalize_cell(value).casefold().replace("ё", "е")
        return re.sub(r"\s+", " ", text).strip()

    @staticmethod
    def _row_has_content(values: tuple[object, ...]) -> bool:
        return any(WorkbookIntakeService._normalize_cell(item) for item in values)

    @staticmethod
    def _get_image_count_by_row(worksheet) -> dict[int, int]:
        counts: dict[int, int] = {}
        for image in getattr(worksheet, "_images", []):
            anchor = getattr(image, "anchor", None)
            marker = getattr(anchor, "_from", None)
            row_index = getattr(marker, "row", None)
            if row_index is None:
                continue
            row_number = int(row_index) + 1
            counts[row_number] = counts.get(row_number, 0) + 1
        return counts

    @staticmethod
    def _get_image_row_bounds(image) -> tuple[int | None, int | None]:
        anchor = getattr(image, "anchor", None)
        start_marker = getattr(anchor, "_from", None)
        end_marker = getattr(anchor, "to", None) or getattr(anchor, "_to", None)
        start_row = getattr(start_marker, "row", None)
        end_row = getattr(end_marker, "row", None)
        start = int(start_row) + 1 if start_row is not None else None
        end = int(end_row) + 1 if end_row is not None else start
        return start, end

    @staticmethod
    def _format_row_list(rows: list[int]) -> str:
        return ", ".join(str(item) for item in sorted(set(rows)))

    def _collect_image_validation_errors(self, worksheet) -> list[str]:
        errors: list[str] = []
        counts = self._get_image_count_by_row(worksheet)
        too_many_rows = [row_number for row_number, image_count in sorted(counts.items()) if image_count > 5]
        if too_many_rows:
            errors.append(
                "Слишком много изображений в строках: "
                f"{self._format_row_list(too_many_rows)}. Допускается максимум 5 изображений на строку."
            )

        cross_row_anchors: list[int] = []
        unknown_anchor_images: list[int] = []
        for index, image in enumerate(getattr(worksheet, "_images", []), start=1):
            start_row, end_row = self._get_image_row_bounds(image)
            if start_row is None:
                unknown_anchor_images.append(index)
                continue
            if end_row is not None and end_row != start_row:
                cross_row_anchors.append(start_row)
        if cross_row_anchors:
            errors.append(
                "Изображения выходят за пределы своей строки в строках: "
                f"{self._format_row_list(cross_row_anchors)}."
            )
        if unknown_anchor_images:
            preview = ", ".join(str(item) for item in unknown_anchor_images[:10])
            tail = f" и еще {len(unknown_anchor_images) - 10}" if len(unknown_anchor_images) > 10 else ""
            errors.append(f"Не удалось определить строку привязки у изображений: {preview}{tail}.")
        return errors

    @staticmethod
    def _resolve_column_index(header_index_map: dict[str, int], candidates: tuple[str, ...]) -> int | None:
        normalized_map = {
            WorkbookIntakeService._normalize_header(header): index for header, index in header_index_map.items()
        }
        for header in candidates:
            if header in header_index_map:
                return header_index_map[header]
            normalized_index = normalized_map.get(WorkbookIntakeService._normalize_header(header))
            if normalized_index is not None:
                return normalized_index
        return None

    def _build_preview_rows(
        self,
        worksheet,
        header_index_map: dict[str, int],
        image_count_by_row: dict[int, int],
    ) -> tuple[int, bool, list[dict[str, object]]]:
        display_col = self._resolve_column_index(header_index_map, DISPLAY_NAME_HEADERS)
        secondary_col = self._resolve_column_index(header_index_map, SECONDARY_TEXT_HEADERS)
        preview_rows: list[dict[str, object]] = []
        total_data_rows = 0

        for row_number in range(2, worksheet.max_row + 1):
            values = tuple(cell.value for cell in worksheet[row_number])
            image_count = image_count_by_row.get(row_number, 0)
            if not self._row_has_content(values) and image_count == 0:
                continue

            total_data_rows += 1
            if len(preview_rows) >= WORKBOOK_PREVIEW_LIMIT:
                continue

            display_name = ""
            secondary_text = ""

            if display_col is not None:
                display_name = self._normalize_cell(worksheet.cell(row_number, display_col).value)
            if secondary_col is not None:
                secondary_text = self._normalize_cell(worksheet.cell(row_number, secondary_col).value)

            if not display_name:
                display_name = next(
                    (self._normalize_cell(value) for value in values if self._normalize_cell(value)),
                    f"Row {row_number}",
                )

            if not secondary_text:
                fallback_values = [
                    self._normalize_cell(value)
                    for value in values
                    if self._normalize_cell(value) and self._normalize_cell(value) != display_name
                ]
                secondary_text = fallback_values[0] if fallback_values else "No secondary text"

            preview_rows.append(
                {
                    "row_number": row_number,
                    "display_name": display_name[:180],
                    "secondary_text": secondary_text[:220],
                    "image_count": image_count,
                    "has_images": image_count > 0,
                }
            )

        return total_data_rows, total_data_rows > WORKBOOK_PREVIEW_LIMIT, preview_rows

    def inspect_workbook(self, workbook_path: str, sheet_name: str | None = None) -> dict[str, object]:
        path = ensure_openpyxl_workbook(Path(workbook_path))
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")

        wb = load_workbook(path, data_only=True)
        try:
            sheet_names = wb.sheetnames
            selected_sheet = sheet_name or (sheet_names[0] if sheet_names else "")

            headers: list[str] = []
            total_data_rows = 0
            preview_truncated = False
            preview_rows: list[dict[str, object]] = []

            matched_required_headers: dict[str, str] = {}
            has_merged_cells = False
            merged_ranges: list[str] = []

            image_validation_errors: list[str] = []

            if selected_sheet and selected_sheet in wb.sheetnames:
                ws = wb[selected_sheet]
                header_index_map: dict[str, int] = {}
                normalized_header_map: dict[str, str] = {}
                for column_index in range(1, ws.max_column + 1):
                    raw_value = ws.cell(1, column_index).value
                    header = str(raw_value).strip() if raw_value is not None else ""
                    if not header:
                        continue
                    headers.append(header)
                    header_index_map[header] = column_index
                    normalized_header_map[self._normalize_header(header)] = header
                merged_ranges = [str(item) for item in getattr(ws.merged_cells, "ranges", [])]
                has_merged_cells = bool(merged_ranges)
                for required_header, aliases in HEADER_NAME_ALIASES.items():
                    for alias in aliases:
                        matched = normalized_header_map.get(alias)
                        if matched:
                            matched_required_headers[required_header] = matched
                            break
                image_count_by_row = self._get_image_count_by_row(ws)
                if not image_count_by_row:
                    image_count_by_row = {
                        row_number: len(images)
                        for row_number, images in collect_xlsx_images_by_row(path, selected_sheet).items()
                    }
                image_validation_errors = self._collect_image_validation_errors(ws)
                total_data_rows, preview_truncated, preview_rows = self._build_preview_rows(
                    ws,
                    header_index_map,
                    image_count_by_row,
                )

            missing_required_headers = [
                header for header in WORKBOOK_REQUIRED_HEADERS if header not in matched_required_headers
            ]
            validation_errors: list[str] = []
            if has_merged_cells:
                preview = ", ".join(merged_ranges[:5])
                tail = f" Еще {len(merged_ranges) - 5}." if len(merged_ranges) > 5 else ""
                validation_errors.append(
                    f"Обнаружены объединенные ячейки: {preview}.{tail}".strip()
                )
            if "Наименование" in missing_required_headers:
                validation_errors.append("Не найден обязательный столбец «Наименование».")
            validation_errors.extend(image_validation_errors)
            return {
                "workbook_path": str(path.resolve()),
                "sheet_names": sheet_names,
                "selected_sheet": selected_sheet,
                "headers": headers,
                "required_headers": WORKBOOK_REQUIRED_HEADERS,
                "optional_headers": WORKBOOK_OPTIONAL_HEADERS,
                "missing_required_headers": missing_required_headers,
                "matched_required_headers": matched_required_headers,
                "has_merged_cells": has_merged_cells,
                "merged_ranges": merged_ranges[:20],
                "image_validation_errors": image_validation_errors,
                "validation_errors": validation_errors,
                "is_workbook_compatible": not validation_errors,
                "total_data_rows": total_data_rows,
                "preview_row_limit": WORKBOOK_PREVIEW_LIMIT,
                "preview_truncated": preview_truncated,
                "preview_rows": preview_rows,
            }
        finally:
            wb.close()

    @staticmethod
    def _default_output_dir(workbook_path: str) -> str:
        workbook = Path(workbook_path).expanduser().resolve()
        return str((workbook.parent / f"{workbook.stem}__agent_cases").resolve())

    def export_case_folders(
        self,
        *,
        workbook_path: str,
        sheet_name: str,
        rows: str,
        output_dir: str | None,
        detect_duplicates: bool,
        header_row: int,
    ) -> dict[str, object]:
        effective_output_dir = output_dir or self._default_output_dir(workbook_path)
        job = self.job_store.create_job(
            job_type="workbook_export",
            module_id="workbook_intake",
            summary=f"Export case folders for {sheet_name} ({rows})",
            command=[],
            payload={
                "workbook_path": workbook_path,
                "sheet_name": sheet_name,
                "rows": rows,
                "output_dir": effective_output_dir,
                "detect_duplicates": detect_duplicates,
                "header_row": header_row,
            },
        )
        self.job_store.update_status(job.job_id, status="running")
        try:
            exported_count = export_workbook_cases(
                workbook_path=workbook_path,
                sheet_name=sheet_name,
                output_dir=effective_output_dir,
                header_row=header_row,
                rows=rows,
                detect_duplicates=detect_duplicates,
            )
            self.source_workbook_service.write_workspace_metadata(
                root_path=Path(effective_output_dir),
                workbook_path=workbook_path,
                sheet_name=sheet_name,
                rows=rows,
            )
            resolved_output_dir = str(Path(effective_output_dir).expanduser().resolve())
            self.runtime_state_service.set_active_case_root_path(resolved_output_dir)
            self.runtime_state_service.ensure_workspace_root(resolved_output_dir)
            self.job_store.update_status(
                job.job_id,
                status="completed",
                output=f"Exported {exported_count} case folder(s) into: {effective_output_dir}",
            )
        except Exception as exc:
            self.job_store.update_status(job.job_id, status="failed", error=str(exc))
        return self.job_store.get_job(job.job_id).__dict__.copy()

    def clear_workbook(self, workbook_path: str) -> dict[str, object]:
        resolved_path = Path(workbook_path).expanduser() if workbook_path.strip() else None
        removed_file = False

        if resolved_path is not None and resolved_path.exists():
            try:
                uploads_root = self.uploads_dir.resolve()
                candidate = resolved_path.resolve()
                if candidate.parent == uploads_root or uploads_root in candidate.parents:
                    candidate.unlink(missing_ok=True)
                    removed_file = True
            except Exception:
                removed_file = False

        return {
            "cleared": True,
            "removed_file": removed_file,
            "workbook_path": workbook_path,
        }
