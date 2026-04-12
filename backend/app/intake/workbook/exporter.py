from __future__ import annotations

import argparse
import hashlib
import json
import os
import posixpath
import re
import subprocess
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

from ...processing.ocr.service import OcrProcessingService


NAME_HEADER = "Наименование"
EXTRA_INFO_HEADER = "Доп информация"
HEADER_ALIASES = {
    NAME_HEADER: {"наименование"},
    EXTRA_INFO_HEADER: {"доп информация", "доп. информация", "доп инфа", "доп. инфа"},
}
REQUIRED_HEADERS = (NAME_HEADER,)
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"


@dataclass(frozen=True)
class ExtractedWorkbookImage:
    path: str
    data: bytes

    def _data(self) -> bytes:
        return self.data


def safe_console_print(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        encoding = getattr(sys.stdout, "encoding", None) or "utf-8"
        safe_text = text.encode(encoding, errors="replace").decode(encoding, errors="replace")
        print(safe_text)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export case folders from workbook rows.")
    parser.add_argument("--workbook", required=True, help="Path to .xlsx or .xls workbook")
    parser.add_argument("--sheet", required=True, help="Worksheet name")
    parser.add_argument(
        "--output-dir",
        help="Target folder for case directories. Defaults to a folder next to the workbook.",
    )
    parser.add_argument("--header-row", type=int, default=1, help="Header row number, default: 1")
    parser.add_argument("--rows", help='Row range, for example: "2-10,15,21-25"')
    parser.add_argument(
        "--long-text-threshold",
        type=int,
        default=400,
        help="Write sidecar .txt when field length is above this threshold",
    )
    parser.add_argument(
        "--detect-duplicates",
        action="store_true",
        help="Detect adjacent duplicate rows by name + image md5 signatures",
    )
    parser.add_argument(
        "--skip-duplicates",
        action="store_true",
        help="Skip exporting rows marked as duplicates. Requires --detect-duplicates.",
    )
    return parser.parse_args()


def default_output_dir_for_workbook(workbook: Path) -> Path:
    return workbook.parent / f"{workbook.stem}__agent_cases"


def ensure_openpyxl_workbook(workbook: Path) -> Path:
    if workbook.suffix.lower() != ".xls":
        return workbook

    converted = workbook.with_name(f"{workbook.stem}__converted.xlsx")
    if converted.exists():
        return converted

    ps_script = r"""
$ErrorActionPreference = 'Stop'
$src = $env:MODULE0_SRC
$dst = $env:MODULE0_DST
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {
  $wb = $excel.Workbooks.Open($src)
  $xlOpenXMLWorkbook = 51
  $wb.SaveAs($dst, $xlOpenXMLWorkbook)
  $wb.Close($false)
} finally {
  if ($wb -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }
  $excel.Quit()
  [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
}
"""
    env = dict(os.environ)
    env["MODULE0_SRC"] = str(workbook)
    env["MODULE0_DST"] = str(converted)
    subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_script],
        check=True,
        env=env,
    )
    if not converted.exists():
        raise SystemExit(f"Failed to convert .xls to .xlsx: {workbook}")
    return converted


def parse_rows(spec: str | None, min_row: int, max_row: int) -> list[int]:
    if not spec:
        return list(range(min_row, max_row + 1))
    result: set[int] = set()
    for chunk in spec.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            left, right = chunk.split("-", 1)
            start = int(left)
            end = int(right)
            if start > end:
                start, end = end, start
            result.update(range(start, end + 1))
        else:
            result.add(int(chunk))
    return [row_number for row_number in sorted(result) if min_row <= row_number <= max_row]


def slugify(value: str, max_length: int = 80) -> str:
    text = re.sub(r"[^\w\s-]", "_", value.strip(), flags=re.UNICODE)
    text = re.sub(r"[-\s]+", "_", text, flags=re.UNICODE).strip("_")
    if not text:
        text = "field"
    return text[:max_length]


def build_header_map(ws, header_row: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for column_index in range(1, ws.max_column + 1):
        value = ws.cell(header_row, column_index).value
        if value is None:
            continue
        header = str(value).strip()
        if header:
            header_map[header] = column_index
    return header_map


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().casefold().replace("ё", "е"))


def resolve_required_headers(header_map: dict[str, int]) -> tuple[dict[str, int], list[str]]:
    normalized_to_index = {normalize_header(header): index for header, index in header_map.items()}
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for required_header in REQUIRED_HEADERS:
        aliases = HEADER_ALIASES.get(required_header, {normalize_header(required_header)})
        found_index = None
        for alias in aliases:
            found_index = normalized_to_index.get(alias)
            if found_index is not None:
                break
        if found_index is None:
            missing.append(required_header)
            continue
        resolved[required_header] = found_index
    optional_aliases = HEADER_ALIASES.get(EXTRA_INFO_HEADER, set())
    for alias in optional_aliases:
        found_index = normalized_to_index.get(alias)
        if found_index is not None:
            resolved[EXTRA_INFO_HEADER] = found_index
            break
    return resolved, missing


def get_image_row(image) -> int | None:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    row_index = getattr(marker, "row", None)
    if row_index is None:
        return None
    return int(row_index) + 1


def get_image_row_bounds(image) -> tuple[int | None, int | None]:
    anchor = getattr(image, "anchor", None)
    start_marker = getattr(anchor, "_from", None)
    end_marker = getattr(anchor, "to", None) or getattr(anchor, "_to", None)
    start_row = getattr(start_marker, "row", None)
    end_row = getattr(end_marker, "row", None)
    start = int(start_row) + 1 if start_row is not None else None
    end = int(end_row) + 1 if end_row is not None else start
    return start, end


def format_row_list(rows: list[int]) -> str:
    return ", ".join(str(item) for item in sorted(set(rows)))


def collect_image_validation_errors(ws) -> list[str]:
    images = list(getattr(ws, "_images", []))
    counts: dict[int, int] = {}
    for image in images:
        row_number = get_image_row(image)
        if row_number is None:
            continue
        counts[row_number] = counts.get(row_number, 0) + 1

    errors: list[str] = []
    too_many_rows = [row_number for row_number, image_count in sorted(counts.items()) if image_count > 5]
    if too_many_rows:
        errors.append(
            "Слишком много изображений в строках: "
            f"{format_row_list(too_many_rows)}. Допускается максимум 5 изображений на строку."
        )

    cross_row_anchors: list[int] = []
    unknown_anchor_images: list[int] = []
    for index, image in enumerate(images, start=1):
        start_row, end_row = get_image_row_bounds(image)
        if start_row is None:
            unknown_anchor_images.append(index)
            continue
        if end_row is not None and end_row != start_row:
            cross_row_anchors.append(start_row)

    if cross_row_anchors:
        errors.append(
            "Изображения выходят за пределы своей строки в строках: "
            f"{format_row_list(cross_row_anchors)}."
        )
    if unknown_anchor_images:
        preview = ", ".join(str(item) for item in unknown_anchor_images[:10])
        tail = f" и еще {len(unknown_anchor_images) - 10}" if len(unknown_anchor_images) > 10 else ""
        errors.append(f"Не удалось определить строку привязки у изображений: {preview}{tail}.")
    return errors


def get_bytes(image) -> bytes:
    ref = getattr(image, "ref", None)
    if ref is None:
        try:
            return image._data()
        except Exception:
            return b""
    return ref.getvalue() if hasattr(ref, "getvalue") else ref


def compute_md5_bytes(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def detect_extension(image) -> str:
    image_path = getattr(image, "path", "") or ""
    suffix = Path(image_path).suffix.lower()
    if suffix in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff"}:
        return suffix
    data = get_bytes(image)
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data.startswith((b"GIF87a", b"GIF89a")):
        return ".gif"
    if data.startswith(b"BM"):
        return ".bmp"
    if data.startswith((b"II*\x00", b"MM\x00*")):
        return ".tif"
    return ".png"


def get_images_by_row(ws) -> dict[int, list[Any]]:
    by_row: dict[int, list[Any]] = {}
    for image in getattr(ws, "_images", []):
        row_number = get_image_row(image)
        if row_number is None:
            continue
        by_row.setdefault(row_number, []).append(image)
    return by_row


def _relationship_source_dir(rels_path: str) -> str:
    if "/_rels/" not in rels_path:
        return posixpath.dirname(rels_path)
    return rels_path.split("/_rels/", 1)[0]


def _resolve_relationship_target(rels_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(_relationship_source_dir(rels_path), target))


def _relationships_path_for(source_path: str) -> str:
    return posixpath.join(posixpath.dirname(source_path), "_rels", f"{posixpath.basename(source_path)}.rels")


def _read_relationships(zf: zipfile.ZipFile, rels_path: str) -> dict[str, tuple[str, str]]:
    if rels_path not in zf.namelist():
        return {}
    root = ElementTree.fromstring(zf.read(rels_path))
    relationships: dict[str, tuple[str, str]] = {}
    for item in root.findall(f"{{{REL_NS}}}Relationship"):
        rel_id = str(item.attrib.get("Id", "")).strip()
        target = str(item.attrib.get("Target", "")).strip()
        rel_type = str(item.attrib.get("Type", "")).strip()
        if not rel_id or not target:
            continue
        relationships[rel_id] = (_resolve_relationship_target(rels_path, target), rel_type)
    return relationships


def _worksheet_path_for_sheet(zf: zipfile.ZipFile, sheet_name: str) -> str | None:
    if "xl/workbook.xml" not in zf.namelist() or "xl/_rels/workbook.xml.rels" not in zf.namelist():
        return None
    workbook_root = ElementTree.fromstring(zf.read("xl/workbook.xml"))
    workbook_rels = _read_relationships(zf, "xl/_rels/workbook.xml.rels")
    rel_attr = f"{{{OFFICE_REL_NS}}}id"
    for sheet in workbook_root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
        if str(sheet.attrib.get("name", "")).strip() != sheet_name:
            continue
        rel_id = str(sheet.attrib.get(rel_attr, "")).strip()
        target, _rel_type = workbook_rels.get(rel_id, ("", ""))
        return target or None
    return None


def collect_xlsx_images_by_row(workbook_path: Path, sheet_name: str) -> dict[int, list[ExtractedWorkbookImage]]:
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"} or not workbook_path.exists():
        return {}
    with zipfile.ZipFile(workbook_path) as zf:
        worksheet_path = _worksheet_path_for_sheet(zf, sheet_name)
        if not worksheet_path:
            return {}
        worksheet_rels = _read_relationships(zf, _relationships_path_for(worksheet_path))
        drawing_paths = [
            target
            for target, rel_type in worksheet_rels.values()
            if rel_type.endswith("/drawing") and target in zf.namelist()
        ]
        by_row: dict[int, list[ExtractedWorkbookImage]] = {}
        for drawing_path in drawing_paths:
            drawing_rels = _read_relationships(zf, _relationships_path_for(drawing_path))
            drawing_root = ElementTree.fromstring(zf.read(drawing_path))
            for anchor_tag in ("twoCellAnchor", "oneCellAnchor"):
                for anchor in drawing_root.findall(f"{{{XDR_NS}}}{anchor_tag}"):
                    row_node = anchor.find(f"{{{XDR_NS}}}from/{{{XDR_NS}}}row")
                    blip = anchor.find(f".//{{{A_NS}}}blip")
                    rel_id = str(blip.attrib.get(f"{{{OFFICE_REL_NS}}}embed", "")).strip() if blip is not None else ""
                    if row_node is None or row_node.text is None or not rel_id:
                        continue
                    try:
                        row_number = int(row_node.text) + 1
                    except ValueError:
                        continue
                    media_path, rel_type = drawing_rels.get(rel_id, ("", ""))
                    if not rel_type.endswith("/image") or media_path not in zf.namelist():
                        continue
                    by_row.setdefault(row_number, []).append(
                        ExtractedWorkbookImage(path=media_path, data=zf.read(media_path))
                    )
        return by_row


def get_row_md5_signatures(images_by_row: dict[int, list[Any]]) -> dict[int, tuple[str, ...]]:
    signatures: dict[int, tuple[str, ...]] = {}
    for row_number, images in images_by_row.items():
        signatures[row_number] = tuple(compute_md5_bytes(get_bytes(image)) for image in images)
    return signatures


def detect_adjacent_duplicates(
    ws,
    row_numbers: list[int],
    name_col_idx: int,
    row_md5_signatures: dict[int, tuple[str, ...]],
) -> dict[int, int]:
    duplicates: dict[int, int] = {}
    previous_name = None
    previous_signature = None
    previous_row = None

    for row_number in row_numbers:
        raw_name = ws.cell(row_number, name_col_idx).value
        name = str(raw_name).strip() if raw_name is not None else ""
        if not name:
            previous_name = None
            previous_signature = None
            previous_row = None
            continue

        current_signature = row_md5_signatures.get(row_number, tuple())
        if name == previous_name and current_signature == previous_signature and previous_row is not None:
            duplicates[row_number] = previous_row
        else:
            previous_name = name
            previous_signature = current_signature
            previous_row = row_number
    return duplicates


def build_same_name_groups(
    ws,
    row_numbers: list[int],
    name_col_idx: int,
) -> tuple[dict[int, str | None], dict[str, list[int]]]:
    rows_by_name: dict[str, list[int]] = {}
    for row_number in row_numbers:
        raw_name = to_text(ws.cell(row_number, name_col_idx).value)
        if not raw_name:
            continue
        rows_by_name.setdefault(raw_name, []).append(row_number)

    group_id_by_row: dict[int, str | None] = {}
    grouped_rows: dict[str, list[int]] = {}
    group_index = 1
    for name, members in rows_by_name.items():
        if len(members) <= 1:
            for row_number in members:
                group_id_by_row[row_number] = None
            continue
        group_id = f"namegrp_{group_index:06d}"
        group_index += 1
        grouped_rows[group_id] = members
        for row_number in members:
            group_id_by_row[row_number] = group_id
    return group_id_by_row, grouped_rows


def format_row_span(rows: list[int]) -> str:
    if not rows:
        return ""
    if len(rows) == 1:
        return str(rows[0])
    return f"{rows[0]}-{rows[-1]}"


def build_case_label(rows: list[int]) -> str:
    row_span = format_row_span(rows)
    if len(rows) <= 1:
        return f"СТРОКА_{row_span}"
    return f"СТРОКИ_{row_span}"


def build_group_members(
    row_numbers: list[int],
    duplicate_map: dict[int, int],
) -> tuple[dict[int, int], dict[int, list[int]]]:
    master_by_row: dict[int, int] = {}
    members_by_master: dict[int, list[int]] = {}
    for row_number in row_numbers:
        master_row = duplicate_map.get(row_number, row_number)
        master_by_row[row_number] = master_row
        members_by_master.setdefault(master_row, []).append(row_number)
    return master_by_row, members_by_master


def write_text_sidecar(case_dir: Path, filename: str, text: str) -> str:
    path = case_dir / filename
    path.write_text(text, encoding="utf-8")
    return filename


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_source_row_payload(
    ws,
    *,
    header_map: dict[str, int],
    header_row: int,
    row_number: int,
    source_file: str,
    sheet_name: str,
    group_members: list[int],
) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    columns: dict[str, str] = {}
    for header, column_index in header_map.items():
        value = to_text(ws.cell(row_number, column_index).value)
        column_letter = ws.cell(header_row, column_index).column_letter
        cells.append(
            {
                "column_index": column_index,
                "column_letter": column_letter,
                "header": header,
                "value": value,
            }
        )
        columns[header] = value

    return {
        "source_type": "workbook",
        "source_file": source_file,
        "sheet_name": sheet_name,
        "header_row": header_row,
        "row_number": row_number,
        "row_span": format_row_span(group_members),
        "source_rows": group_members,
        "cells": cells,
        "columns": columns,
    }


def build_case_payload(
    *,
    case_id: str,
    case_dir_name: str,
    source_file: str,
    sheet_name: str,
    row_number: int,
    group_members: list[int],
    raw_name: str,
    extra_info: str,
    name_file: str | None,
    extra_info_file: str | None,
    image_files: list[str],
    duplicate_of: int | None,
    master_row: int,
    same_name_group_id: str | None,
    same_name_group_size: int,
) -> dict[str, Any]:
    row_span = format_row_span(group_members)
    duplicate_group_id = f"dupgrp_{master_row:06d}" if len(group_members) > 1 else None
    master_case_id = build_case_label(group_members)
    case_role = "master"
    image_names = [Path(item).name for item in image_files]
    image_count = len(image_files)

    payload = {
        "case_id": case_id,
        "case_dir_name": case_dir_name,
        "source_file": source_file,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "source_rows": group_members,
        "row_span": row_span,
        "raw_name": raw_name,
        "extra_info": extra_info,
        "name_file": name_file,
        "extra_info_file": extra_info_file,
        "image_files": image_files,
        "case_role": case_role,
        "is_duplicate": len(group_members) > 1,
        "duplicate_of_row": duplicate_of,
        "master_row_number": master_row,
        "master_case_id": master_case_id,
        "duplicate_group_id": duplicate_group_id,
        "duplicate_group_size": len(group_members),
        "duplicate_rows": [row for row in group_members if row != master_row],
        "same_name_group_id": same_name_group_id,
        "same_name_group_size": same_name_group_size,
        "status": "prepared",
        "source": {
            "source_type": "workbook",
            "source_file": source_file,
            "sheet_name": sheet_name,
            "row_number": row_number,
            "row_span": row_span,
            "source_rows": group_members,
        },
        "product": {
            "raw_name": raw_name,
            "extra_info": extra_info,
        },
        "media": {
            "image_files": image_names,
            "image_count": image_count,
        },
        "groups": {
            "duplicate_group_id": duplicate_group_id,
            "duplicate_group_size": len(group_members),
            "same_name_group_id": same_name_group_id,
            "same_name_group_size": same_name_group_size,
        },
        "status_view": {
            "review_status": "pending",
            "prefetch_status": "idle",
            "work_status": "prepared",
            "last_completed_stage": "workbook_intake",
        },
        "links": {
            "source_row": "source_row.json",
            "work_status": "work/status.json",
            "ocr": "work/ocr.json",
            "tnved": "work/tnved.json",
            "verification": "work/verification.json",
            "tnved_vbd": "work/tnved_vbd.json",
            "enrichment": "work/enrichment.json",
            "calculations": "work/calculations.json",
            "questions": "work/questions.json",
            "pipeline_result": "result/pipeline_result.json",
            "ui_response": "result/ui_response.json",
            "export": "result/export.json",
        },
    }
    return payload


def write_stage_placeholders(case_dir: Path, *, case_id: str, row_number: int, raw_name: str) -> None:
    work_dir = case_dir / "work"
    result_dir = case_dir / "result"
    work_dir.mkdir(parents=True, exist_ok=True)
    result_dir.mkdir(parents=True, exist_ok=True)

    status_payload = {
        "case_id": case_id,
        "row_number": row_number,
        "current_stage": "workbook_intake",
        "last_completed_stage": "workbook_intake",
        "failed_stage": "",
        "status": "prepared",
        "error_text": "",
        "user_verified": False,
    }
    ocr_seed = OcrProcessingService().build_seed_payload(
        case_payload={
            "raw_name": raw_name,
            "extra_info": "",
            "product": {"raw_name": raw_name, "extra_info": ""},
        },
        source_row_payload=None,
    )

    placeholders: dict[Path, dict[str, Any]] = {
        work_dir / "status.json": status_payload,
        work_dir / "ocr.json": ocr_seed,
        work_dir / "tnved.json": {
            "status": "pending",
            "long_description": "",
            "selected_code": "",
            "selected_description": "",
            "selection_rationale": "",
            "confidence_percent": None,
            "candidates": [],
        },
        work_dir / "verification.json": {
            "status": "pending",
            "validation_status": "",
            "semantic_status": "",
            "ifcg_status": "",
            "notes": [],
        },
        work_dir / "tnved_vbd.json": {
            "status": "pending",
            "verification_status": "pending",
            "selected_code": "",
            "summary": "",
            "note": "",
            "product_facts": [],
            "reference_hits": [],
            "example_hits": [],
            "alternative_codes": [],
            "warnings": [],
            "index_status": "pending",
            "trace": {},
        },
        work_dir / "enrichment.json": {
            "status": "pending",
            "ifcg_discovery": {"status": "pending", "summary": "", "suggested_codes": [], "suggested_groups": []},
            "tnved_vbd": {
                "status": "pending",
                "verification_status": "pending",
                "selected_code": "",
                "summary": "",
                "note": "",
                "product_facts": [],
                "reference_hits": [],
                "example_hits": [],
                "alternative_codes": [],
                "warnings": [],
                "index_status": "pending",
                "trace": {},
            },
            "ifcg_verification": {"status": "pending"},
            "its": {"status": "pending"},
            "sigma": {"status": "pending"},
        },
        work_dir / "calculations.json": {
            "status": "pending",
            "customs": {"status": "pending"},
            "stp": {"status": "pending", "value": None, "currency": "USD"},
            "eco_fee": {"status": "pending", "preview": "", "matches": []},
        },
        work_dir / "questions.json": {
            "status": "pending",
            "count": 0,
            "top": [],
            "short": [],
            "answers": [],
            "final_code": "",
            "final_status": "pending",
        },
        result_dir / "ui_response.json": {
            "workspace": {"case_id": case_id, "row_number": row_number, "title": raw_name},
            "input": {},
            "questions": {"status": "pending", "count": 0, "top": [], "short": [], "answers": []},
            "decision": {},
            "enrichment": {},
            "calculations": {},
        },
        result_dir / "pipeline_result.json": {
            "schema_version": "case_pipeline.v1",
            "status": "pending",
            "input": {},
            "questions": {"status": "pending", "count": 0, "top": [], "short": [], "answers": []},
            "decision": {},
            "enrichment": {
                "ifcg_discovery": None,
                "tnved_vbd": None,
                "ifcg_verification": None,
                "its": None,
                "sigma": None,
            },
            "calculations": {
                "customs": None,
                "stp": None,
                "eco_fee": None,
            },
            "stages": {
                "ifcg_discovery": None,
                "tnved": None,
                "semantic": None,
                "verification": None,
                "tnved_vbd": None,
                "ifcg_verification": None,
            },
            "warnings": [],
            "error_text": "",
            "trace": {},
        },
        result_dir / "export.json": {
            "case_id": case_id,
            "selected_code": "",
            "selected_description": "",
            "operator_comment": "",
        },
    }
    for path, payload in placeholders.items():
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_case(
    ws,
    header_map: dict[str, int],
    header_row: int,
    row_number: int,
    name_col_idx: int,
    extra_info_col_idx: int | None,
    images_by_row: dict[int, list[Any]],
    output_dir: Path,
    long_text_threshold: int,
    source_file: str,
    sheet_name: str,
    duplicate_of: int | None,
    master_row: int,
    group_members: list[int],
    same_name_group_id: str | None,
    same_name_group_size: int,
) -> dict[str, Any]:
    row_span = format_row_span(group_members)
    case_id = build_case_label(group_members)
    case_dir = output_dir / case_id
    images_dir = case_dir / "images"
    case_dir.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)

    raw_name = to_text(ws.cell(row_number, name_col_idx).value)
    extra_info = to_text(ws.cell(row_number, extra_info_col_idx).value) if extra_info_col_idx else ""

    name_file = None
    extra_info_file = None
    if raw_name and (len(raw_name) > long_text_threshold or "\n" in raw_name):
        name_file = write_text_sidecar(case_dir, "name.txt", raw_name)
    if extra_info and (len(extra_info) > long_text_threshold or "\n" in extra_info):
        extra_info_file = write_text_sidecar(case_dir, "extra_info.txt", extra_info)

    image_files: list[str] = []
    for image_index, image in enumerate(images_by_row.get(row_number, []), start=1):
        extension = detect_extension(image)
        filename = f"image_{image_index:02d}{extension}"
        target = images_dir / filename
        target.write_bytes(get_bytes(image))
        image_files.append(f"images/{filename}")

    case_payload = build_case_payload(
        case_id=case_id,
        case_dir_name=case_dir.name,
        source_file=source_file,
        sheet_name=sheet_name,
        row_number=row_number,
        group_members=group_members,
        raw_name=raw_name,
        extra_info=extra_info,
        name_file=name_file,
        extra_info_file=extra_info_file,
        image_files=image_files,
        duplicate_of=duplicate_of,
        master_row=master_row,
        same_name_group_id=same_name_group_id,
        same_name_group_size=same_name_group_size,
    )
    (case_dir / "case.json").write_text(
        json.dumps(case_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    source_row_payload = build_source_row_payload(
        ws,
        header_map=header_map,
        header_row=header_row,
        row_number=row_number,
        source_file=source_file,
        sheet_name=sheet_name,
        group_members=group_members,
    )
    (case_dir / "source_row.json").write_text(
        json.dumps(source_row_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_stage_placeholders(case_dir, case_id=case_id, row_number=row_number, raw_name=raw_name)
    OcrProcessingService().seed_case_file(
        case_dir=case_dir,
        case_payload=case_payload,
        source_row_payload=source_row_payload,
    )
    return case_payload


def export_workbook_cases(
    *,
    workbook_path: str,
    sheet_name: str,
    output_dir: str | None = None,
    header_row: int = 1,
    rows: str | None = None,
    long_text_threshold: int = 400,
    detect_duplicates: bool = True,
    skip_duplicates: bool = False,
) -> int:
    workbook = Path(workbook_path).expanduser().resolve()
    workbook_for_openpyxl = ensure_openpyxl_workbook(workbook)
    target_output_dir = (
        Path(output_dir).expanduser().resolve()
        if output_dir
        else default_output_dir_for_workbook(workbook)
    )
    if skip_duplicates and not detect_duplicates:
        raise ValueError("--skip_duplicates requires detect_duplicates=True")

    wb = load_workbook(workbook_for_openpyxl, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        if getattr(ws.merged_cells, "ranges", None):
            merged_preview = ", ".join(str(item) for item in list(ws.merged_cells.ranges)[:8])
            raise ValueError(f"Workbook contains merged cells and cannot be processed: {merged_preview}")
        image_validation_errors = collect_image_validation_errors(ws)
        if image_validation_errors:
            raise ValueError(" ".join(image_validation_errors))

        header_map = build_header_map(ws, header_row)
        resolved_headers, missing = resolve_required_headers(header_map)
        if missing:
            raise ValueError(f"Missing headers: {', '.join(missing)}")

        name_col_idx = resolved_headers[NAME_HEADER]
        extra_info_col_idx = resolved_headers.get(EXTRA_INFO_HEADER)
        target_output_dir.mkdir(parents=True, exist_ok=True)

        row_numbers = parse_rows(rows, header_row + 1, ws.max_row)
        images_by_row = get_images_by_row(ws)
        if not images_by_row:
            images_by_row = collect_xlsx_images_by_row(workbook_for_openpyxl, sheet_name)
        row_md5_signatures = get_row_md5_signatures(images_by_row) if detect_duplicates else {}
        duplicate_map = (
            detect_adjacent_duplicates(ws, row_numbers, name_col_idx, row_md5_signatures)
            if detect_duplicates
            else {}
        )
        master_by_row, members_by_master = build_group_members(row_numbers, duplicate_map)
        same_name_group_id_by_row, same_name_group_members = build_same_name_groups(
            ws, row_numbers, name_col_idx
        )

        manifest: list[dict[str, Any]] = []
        row_index: list[dict[str, Any]] = []
        processed_masters: set[int] = set()
        for row_number in row_numbers:
            master_row = master_by_row[row_number]
            group_members = members_by_master[master_row]
            duplicate_of = duplicate_map.get(row_number)
            row_index.append(
                {
                    "row_number": row_number,
                    "master_row_number": master_row,
                    "case_id": build_case_label(group_members),
                    "row_span": format_row_span(group_members),
                    "is_duplicate": row_number != master_row,
                    "duplicate_of_row": duplicate_of,
                    "duplicate_group_id": f"dupgrp_{master_row:06d}" if len(group_members) > 1 else None,
                    "same_name_group_id": same_name_group_id_by_row.get(row_number),
                }
            )
            if master_row in processed_masters:
                continue
            processed_masters.add(master_row)
            if skip_duplicates and len(group_members) > 1:
                # Still export one grouped master folder.
                pass

            manifest.append(
                write_case(
                    ws=ws,
                    header_map=header_map,
                    header_row=header_row,
                    row_number=master_row,
                    name_col_idx=name_col_idx,
                    extra_info_col_idx=extra_info_col_idx,
                    images_by_row=images_by_row,
                    output_dir=target_output_dir,
                    long_text_threshold=long_text_threshold,
                    source_file=workbook.name,
                    sheet_name=sheet_name,
                    duplicate_of=None,
                    master_row=master_row,
                    group_members=group_members,
                    same_name_group_id=same_name_group_id_by_row.get(master_row),
                    same_name_group_size=len(
                        same_name_group_members.get(
                            same_name_group_id_by_row.get(master_row) or "", []
                        )
                    ),
                )
            )

        manifest_path = target_output_dir / "manifest.json"
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        review_queue = []
        duplicate_groups = []
        same_name_groups = []
        manifest_by_row = {item["row_number"]: item for item in manifest if "row_number" in item}
        for master_row, member_rows in sorted(members_by_master.items()):
            master_item = manifest_by_row.get(master_row)
            if not master_item:
                continue
            if len(member_rows) > 1:
                duplicate_groups.append(
                    {
                        "duplicate_group_id": f"dupgrp_{master_row:06d}",
                        "master_case_id": master_item["case_id"],
                        "master_case_dir_name": master_item["case_dir_name"],
                        "master_row_number": master_row,
                    "member_case_ids": [master_item["case_id"]],
                        "member_case_dir_names": [master_item["case_dir_name"]],
                        "member_rows": member_rows,
                        "duplicate_rows": [row for row in member_rows if row != master_row],
                        "total_cases": len(member_rows),
                    }
                )
            review_queue.append(
                {
                    "case_id": master_item["case_id"],
                    "case_dir_name": master_item["case_dir_name"],
                    "row_number": master_row,
                    "source_rows": member_rows,
                    "row_span": format_row_span(member_rows),
                    "raw_name": master_item.get("raw_name", ""),
                    "image_files": master_item.get("image_files", []),
                    "duplicate_group_id": f"dupgrp_{master_row:06d}" if len(member_rows) > 1 else None,
                    "duplicate_group_size": len(member_rows),
                    "duplicate_count": max(len(member_rows) - 1, 0),
                    "same_name_group_id": master_item.get("same_name_group_id"),
                    "same_name_group_size": master_item.get("same_name_group_size", 0),
                    "status": "ready_for_review",
                }
            )

        for group_id, member_rows in sorted(same_name_group_members.items()):
            if not member_rows:
                continue
            representative = manifest_by_row.get(member_rows[0], {})
            grouped_case_ids = []
            for row in member_rows:
                case_label = build_case_label(members_by_master[master_by_row[row]])
                if case_label not in grouped_case_ids:
                    grouped_case_ids.append(case_label)
            same_name_groups.append(
                {
                    "same_name_group_id": group_id,
                    "raw_name": representative.get("raw_name", ""),
                    "member_case_ids": grouped_case_ids,
                    "member_rows": member_rows,
                    "total_cases": len(member_rows),
                }
            )

        (target_output_dir / "review_queue.json").write_text(
            json.dumps(review_queue, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (target_output_dir / "duplicate_groups.json").write_text(
            json.dumps(duplicate_groups, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (target_output_dir / "same_name_groups.json").write_text(
            json.dumps(same_name_groups, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (target_output_dir / "row_index.json").write_text(
            json.dumps(row_index, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return len(manifest)
    finally:
        wb.close()


def main() -> int:
    args = parse_args()
    exported_count = export_workbook_cases(
        workbook_path=args.workbook,
        sheet_name=args.sheet,
        output_dir=args.output_dir,
        header_row=args.header_row,
        rows=args.rows,
        long_text_threshold=args.long_text_threshold,
        detect_duplicates=args.detect_duplicates,
        skip_duplicates=args.skip_duplicates,
    )
    workbook = Path(args.workbook).expanduser().resolve()
    target_output_dir = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else default_output_dir_for_workbook(workbook)
    )
    safe_console_print(f"Exported {exported_count} case folder(s) into: {target_output_dir}")
    safe_console_print(f"Manifest: {target_output_dir / 'manifest.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
