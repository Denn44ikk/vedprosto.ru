from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ...config import AppSettings


WORKSPACE_SOURCE_META_FILE = "workspace_source.json"
DEFAULT_HEADER_ROW = 1
_FALLBACK_SOURCE_HEADERS = {
    "source_file": "Источник файла",
    "sheet_name": "Лист",
    "row_span": "Строки",
    "raw_name": "Наименование",
    "extra_info": "Доп инфа",
    "duplicate_group_size": "Duplicate group",
    "same_name_group_size": "Same-name group",
}
_HEADER_PRIORITY_GROUPS: tuple[tuple[str, ...], ...] = (
    ("наименование", "название", "product", "name", "товар"),
    ("опис", "description", "desc"),
    ("доп", "note", "comment", "备注"),
    ("артик", "article", "sku"),
    ("бренд", "brand", "tm", "торгов"),
    ("модель", "model"),
    ("материал", "состав", "composition", "material"),
    ("назнач", "использ", "function", "usage"),
    ("страна", "country"),
    ("производ", "manufacturer", "factory", "supplier"),
    ("кол", "qty", "quantity", "数量"),
    ("цена", "price", "стоим", "usd", "cny"),
    ("вес", "нетто", "брутто", "kg", "кг"),
    ("упаков", "package", "tare"),
    ("фото", "image", "img"),
)


class SourceWorkbookService:
    def __init__(self, *, settings: AppSettings) -> None:
        self.settings = settings
        self._resolved_cache: dict[tuple[str, str], str | None] = {}

    @staticmethod
    def _normalize(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _meta_path(root_path: Path) -> Path:
        return root_path / WORKSPACE_SOURCE_META_FILE

    def write_workspace_metadata(
        self,
        *,
        root_path: str | Path,
        workbook_path: str,
        sheet_name: str,
        rows: str,
    ) -> None:
        target_root = Path(root_path).expanduser().resolve()
        if not target_root.exists():
            return
        payload = {
            "workbook_path": str(Path(workbook_path).expanduser().resolve()),
            "sheet_name": sheet_name,
            "rows": rows,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        self._meta_path(target_root).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def read_workspace_metadata(self, root_path: Path) -> dict[str, object]:
        meta_path = self._meta_path(root_path)
        if not meta_path.exists():
            return {}
        try:
            payload = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
        return payload if isinstance(payload, dict) else {}

    def _search_by_filename(self, *, source_file: str) -> Path | None:
        cache_key = ("project", source_file)
        if cache_key in self._resolved_cache:
            cached = self._resolved_cache[cache_key]
            return Path(cached) if cached else None

        resolved: Path | None = None
        try:
            for candidate in self.settings.project_root.rglob(source_file):
                if candidate.is_file():
                    resolved = candidate.resolve()
                    break
        except Exception:
            resolved = None

        self._resolved_cache[cache_key] = str(resolved) if resolved else None
        return resolved

    def resolve_source_workbook_path(self, *, root_path: Path, case_payload: dict) -> Path | None:
        source_file = self._normalize(case_payload.get("source_file"))
        if not source_file:
            return None

        cache_key = (str(root_path.resolve()), source_file)
        if cache_key in self._resolved_cache:
            cached = self._resolved_cache[cache_key]
            return Path(cached) if cached else None

        metadata = self.read_workspace_metadata(root_path)
        metadata_path = self._normalize(metadata.get("workbook_path"))
        if metadata_path:
            candidate = Path(metadata_path).expanduser()
            if candidate.exists():
                resolved = candidate.resolve()
                self._resolved_cache[cache_key] = str(resolved)
                return resolved

        raw_candidate = Path(source_file).expanduser()
        direct_candidates = []
        if raw_candidate.is_absolute():
            direct_candidates.append(raw_candidate)
        direct_candidates.extend(
            [
                root_path.parent / source_file,
                root_path / source_file,
                root_path.parent.parent / source_file,
            ]
        )
        for candidate in direct_candidates:
            if candidate.exists() and candidate.is_file():
                resolved = candidate.resolve()
                self._resolved_cache[cache_key] = str(resolved)
                return resolved

        resolved = self._search_by_filename(source_file=source_file)
        self._resolved_cache[cache_key] = str(resolved) if resolved else None
        return resolved

    @staticmethod
    def _field_priority(label: str) -> tuple[int, str]:
        normalized = label.strip().lower()
        for index, tokens in enumerate(_HEADER_PRIORITY_GROUPS):
            if any(token in normalized for token in tokens):
                return index, normalized
        return len(_HEADER_PRIORITY_GROUPS), normalized

    @classmethod
    def _sorted_fields(cls, fields: list[dict[str, object]]) -> list[dict[str, object]]:
        return sorted(
            fields,
            key=lambda item: cls._field_priority(str(item.get("label", ""))),
        )

    def _build_fallback_table(self, *, case_payload: dict) -> dict[str, object]:
        fields = []
        for key, label in _FALLBACK_SOURCE_HEADERS.items():
            value = self._normalize(case_payload.get(key))
            if not value:
                continue
            fields.append({"label": label, "values": [value]})

        return {
            "status": "fallback",
            "workbook_name": self._normalize(case_payload.get("source_file")) or "—",
            "workbook_path": "",
            "sheet_name": self._normalize(case_payload.get("sheet_name")) or "—",
            "row_labels": [f"стр. {int(case_payload.get('row_number', 0) or 0)}"],
            "note": "Исходный Excel не найден. Пока показаны поля из case.json.",
            "fields": self._sorted_fields(fields),
        }

    def build_source_table(self, *, root_path: Path, case_payload: dict) -> dict[str, object]:
        row_numbers = [
            int(item)
            for item in case_payload.get("source_rows", [])
            if isinstance(item, int)
        ]
        if not row_numbers:
            row_number = int(case_payload.get("row_number", 0) or 0)
            if row_number:
                row_numbers = [row_number]

        workbook_path = self.resolve_source_workbook_path(root_path=root_path, case_payload=case_payload)
        if workbook_path is None:
            return self._build_fallback_table(case_payload=case_payload)

        sheet_name = self._normalize(case_payload.get("sheet_name"))
        workbook_name = workbook_path.name

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                return {
                    "status": "sheet_missing",
                    "workbook_name": workbook_name,
                    "workbook_path": str(workbook_path),
                    "sheet_name": sheet_name or "—",
                    "row_labels": [f"стр. {row}" for row in row_numbers],
                    "note": f"Лист {sheet_name or '—'} не найден в исходном Excel.",
                    "fields": [],
                }

            worksheet = workbook[sheet_name]
            fields: list[dict[str, object]] = []
            for column_index in range(1, worksheet.max_column + 1):
                header = self._normalize(worksheet.cell(DEFAULT_HEADER_ROW, column_index).value)
                row_values = [
                    self._normalize(worksheet.cell(row_number, column_index).value)
                    for row_number in row_numbers
                ]
                if not header and not any(row_values):
                    continue

                label = header or f"Колонка {get_column_letter(column_index)}"
                fields.append(
                    {
                        "label": label,
                        "values": [value or "—" for value in row_values],
                    }
                )

            if not fields:
                return {
                    "status": "empty",
                    "workbook_name": workbook_name,
                    "workbook_path": str(workbook_path),
                    "sheet_name": sheet_name or "—",
                    "row_labels": [f"стр. {row}" for row in row_numbers],
                    "note": "В выбранных строках нет читаемых значений.",
                    "fields": [],
                }

            return {
                "status": "ready",
                "workbook_name": workbook_name,
                "workbook_path": str(workbook_path),
                "sheet_name": sheet_name or "—",
                "row_labels": [f"стр. {row}" for row in row_numbers],
                "note": f"{workbook_name} · {sheet_name or '—'} · {', '.join(f'стр. {row}' for row in row_numbers)}",
                "fields": self._sorted_fields(fields),
            }
        finally:
            workbook.close()
